import os
import io
import time
import datetime
import re
from pathlib import Path
from urllib import request as urlrequest
from typing import Dict, Any, Optional, List, Tuple

# PDF Generation imports
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Excel Generation imports
try:
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from calculators import calculate_lenticule_thickness_smile, alpins_vector_analysis

def _get_p(db: Dict[str, Any], pid: Any) -> Dict[str, Any]:
    for p in db.get("patients", []):
        try:
            if str(p.get("patient_id")) == str(pid): return p
        except Exception: continue
    return {}

def _get_m(db: Dict[str, Any], pid: Any) -> Tuple[Optional[str], Dict[str, Any]]:
    v = db.get("visits", {}).get(str(pid)) or {}
    vid = v.get("visit_id")
    if not vid: return None, {}
    m = db.get("meas", {}).get(vid) or {}
    return str(vid), (m if isinstance(m, dict) else {})

def _register_cyrillic_font(tmp_dir: Path) -> str:
    """Attempts to register a Cyrillic font for ReportLab."""
    font_path = tmp_dir / "Roboto-Regular.ttf"
    
    if not font_path.exists():
        try:
            url = "https://raw.githubusercontent.com/googlefonts/roboto/main/src/hinted/Roboto-Regular.ttf"
            req = urlrequest.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urlrequest.urlopen(req, timeout=15) as resp, open(font_path, 'wb') as f:
                f.write(resp.read())
        except Exception as e:
            print(f"[WARN] Failed to download font: {e}")
            
    candidates = [
        str(font_path),
        "DejaVuSans.ttf",
        "Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf"
    ]
    for path in candidates:
        try:
            if os.path.exists(path) or (not os.path.isabs(path) and os.path.exists(path)):
                pdfmetrics.registerFont(TTFont('Regular', path))
                return 'Regular'
        except Exception:
            continue
    return 'Helvetica' # Fallback

def generate_day_schedule_pdf(date_str: str, db: Dict[str, Any], tmp_dir: Path) -> Optional[str]:
    if not PDF_AVAILABLE:
        return None
    
    forms = db.get("forms", {})
    visits = db.get("visits", {})
    
    pats = []
    for pid, f in forms.items():
        if f.get("op_date") != date_str: continue
        v = visits.get(pid)
        if not v or v.get("status") != "approved": continue
        p = _get_p(db, pid)
        if p:
            op_time = str(f.get("op_time") or "23:59")
            pats.append((pid, p, op_time))
            
    if not pats: return None

    out_path = tmp_dir / f"schedule_{date_str}.pdf"
    c = canvas.Canvas(str(out_path), pagesize=A4)
    width, height = A4
    font_name = _register_cyrillic_font(tmp_dir)
    
    y = height - 50
    c.setFont(font_name, 16)
    c.drawString(50, y, f"Операционный список: {date_str}")
    y -= 40
    
    pats.sort(key=lambda x: x[2])
    
    for pid, p, op_time in pats:
        if y < 120:
            c.showPage()
            y = height - 50
            c.setFont(font_name, 16)
            c.drawString(50, y, f"Операционный список: {date_str} (продолжение)")
            y -= 40
            
        vid, m = _get_m(db, pid)
        plan = m.get("surgery_plan", {})
        flap = plan.get("flap", {})
        iol = m.get("iol_calc", {})
        is_cataract = bool(iol)
        
        name = p.get("name", "-")
        f = forms.get(pid, {})
        prim = f.get("primary", {})
        age = str(prim.get("age", "")).strip()
        age_label = f", {age} лет" if age else ""
        time_label = f"[{op_time}] " if op_time != "23:59" else ""

        if is_cataract:
            lb = iol.get("lens_barrett") or iol.get("lens_name", "-")
            ab = iol.get("a_const_barrett") or iol.get("a_const", "-")
            lk = iol.get("lens_kane") or iol.get("lens_name", "-")
            ak = iol.get("a_const_kane", "-")
            
            if lb == lk and ab != "-" and ak != "-":
                a_str = f"{lb} (A: {ab}/{ak})"
            elif lb == lk:
                a_str = f"{lb} (A: {ab if ab != '-' else ak})"
            else:
                a_str = f"B:{lb}({ab}) | K:{lk}({ak})"
            
            c.setFont(font_name, 12)
            c.drawString(50, y, f"{time_label}{name}{age_label} (ID: {pid}) - ИОЛ: {a_str}")
            y -= 20
            
            op_eye = str(f.get("op_eye") or "ou").lower()
            for side, label in [("od", "OD"), ("os", "OS")]:
                if op_eye in ("od", "os") and side != op_eye:
                    continue
                al = m.get("axial_length", {}).get(side, {}).get("value")
                ker = m.get("keratometry", {}).get(side, {})
                k1 = ker.get("k1"); k2 = ker.get("k2")
                
                if not al and not k1: continue
                
                c.setFont(font_name, 10)
                c.drawString(50, y, f"{label}: AL {al or '-'} | K1 {k1 or '-'} | K2 {k2 or '-'}")
                y -= 12
                
                if al and k1 and k2:
                    try:
                        sel_iol = iol.get("selected_iol", {}).get(side, {})
                        if sel_iol and sel_iol.get("power"):
                            c.setFont(font_name, 10)
                            c.setFillColorRGB(0.1, 0.6, 0.1) # Greenish
                            model_str = sel_iol.get("model", "")
                            if model_str:
                                c.drawString(70, y, f"ПЛАН ИОЛ: {model_str} | {sel_iol['power']} D (Target: {sel_iol.get('target', '-')})")
                            else:
                                c.drawString(70, y, f"ПЛАН ИОЛ: {sel_iol['power']} D (Target: {sel_iol.get('target', '-')})")
                            c.setFillColorRGB(0, 0, 0)
                            y -= 12
                            
                        # Если есть торическая схема Kane, рисуем её справа
                        t_path = iol.get("kane", {}).get(side, {}).get("toric_img_path")
                        if t_path and os.path.exists(t_path):
                            try:
                                c.drawImage(t_path, width - 110, y - 50, width=60, height=60, preserveAspectRatio=True)
                            except: pass
                            
                        tables_to_print = []
                        if iol.get("barrett", {}).get(side, {}).get("table"):
                            tables_to_print.append(("Barrett:", iol["barrett"][side]["table"]))
                        if iol.get("kane", {}).get(side, {}).get("table"):
                            tables_to_print.append(("Kane:", iol["kane"][side]["table"]))
                        if iol.get("escrs", {}).get(side):
                            for f_name, f_data in iol["escrs"][side].items():
                                if f_name == "error": continue
                                if f_name.lower() in ("kane", "barrett"): continue
                                if isinstance(f_data, dict) and "table" in f_data:
                                    tables_to_print.append((f"{f_name}:", f_data["table"]))
                            
                        for prefix, tbl in tables_to_print:
                            if tbl:
                                tbl_str = f"{prefix} "
                                # Ограничиваем длину таблицы, если рисуем картинку справа
                                max_cols = 6 if (t_path and os.path.exists(t_path)) else 9
                                for row in tbl[:max_cols]:
                                    tbl_str += f"{row['power']:.1f}({row['ref']:+.2f})   "
                                c.setFont(font_name, 9)
                                c.setFillColorRGB(0.3, 0.3, 0.3)
                                c.drawString(70, y, tbl_str)
                                c.setFillColorRGB(0, 0, 0)
                                y -= 12
                    except: pass
                y -= 5
        else:
            laser_type = plan.get("laser_type")
            laser_map = {
                "ex500": "Alcon EX500 (WFO)", "visx_s4ir": "VISX Star S4 IR",
                "visumax_800": "VisuMax 800 (SMILE Pro)", "visumax_500": "VisuMax 500 (SMILE)",
                "smartsight": "SmartSight (Schwind)", "silk": "SILK (J&J Elita)", "mel90": "Zeiss MEL 90"
            }
            laser_display_name = laser_map.get(laser_type, laser_type) if laser_type else "Не указан"
            c.setFont(font_name, 12)
            c.drawString(50, y, f"{time_label}{name}{age_label} (ID: {pid}) - Лазер: {laser_display_name}")
            y -= 20
            
            c.setFont(font_name, 10)
            
            def _get_data(side):
                rx = plan.get(side, {})
                if not isinstance(rx, dict): rx = {}
                sph = rx.get("sph"); cyl = rx.get("cyl"); ax = rx.get("axis")
                f_ = flap.get(side, {})
                if not isinstance(f_, dict): f_ = {}
                th = f_.get("thickness_um"); dm = f_.get("diameter_mm")
                if th is None: th = flap.get("thickness_um")
                if dm is None: dm = flap.get("diameter_mm")
                try: ref_str = f"{float(sph):+.2f} {float(cyl):+.2f} x {int(float(ax))}" if (sph not in (None, "") and cyl not in (None, "") and ax not in (None, "")) else "-"
                except: ref_str = "-"
                try: flap_str = f"{int(float(th))} µm / {float(dm):.1f} mm" if (th not in (None, "") and dm not in (None, "")) else "-"
                except: flap_str = "-"
                return ref_str, flap_str

            is_smile = str(laser_type or "").startswith("visumax") or laser_type in ("smartsight", "silk")
            ref_label = "Femto" if is_smile else "Excimer"
            od_ref, od_flap = _get_data("od")
            os_ref, os_flap = _get_data("os")
            
            op_eye = str(f.get("op_eye") or "ou").lower()
            
            if op_eye in ("ou", "od"):
                c.drawString(50, y, "OD")
                c.drawString(80, y, f"{ref_label}: {od_ref}")
                if not is_smile and od_flap != "-":
                    c.drawString(300, y, f"Femto: {od_flap}")
                elif is_smile:
                    cap_th = plan.get("cap", {}).get("od", {}).get("thickness_um", "-")
                    min_th = plan.get("smile", {}).get("od", {}).get("min_thickness_um", "10" if laser_type == "visumax_800" else "15")
                    len_th = "-"
                    try:
                        rx = plan.get("od", {})
                        if isinstance(rx, dict) and rx.get("sph") not in (None, ""):
                            oz = float(rx.get("optical_zone") or plan.get("optical_zone_mm") or 6.5)
                            depth = calculate_lenticule_thickness_smile(float(rx["sph"]), float(rx.get("cyl", 0)), oz)
                            len_th = str(int(depth + (float(min_th) if min_th not in (None, "") else 15.0)))
                    except: pass
                    c.drawString(300, y, f"Cap: {cap_th} / Min: {min_th} / Len: {len_th} µm")
                y -= 15
            
            if op_eye in ("ou", "os"):
                c.drawString(50, y, "OS")
                c.drawString(80, y, f"{ref_label}: {os_ref}")
                if not is_smile and os_flap != "-":
                    c.drawString(300, y, f"Femto: {os_flap}")
                elif is_smile:
                    cap_th = plan.get("cap", {}).get("os", {}).get("thickness_um", "-")
                    min_th = plan.get("smile", {}).get("os", {}).get("min_thickness_um", "10" if laser_type == "visumax_800" else "15")
                    len_th = "-"
                    try:
                        rx = plan.get("os", {})
                        if isinstance(rx, dict) and rx.get("sph") not in (None, ""):
                            oz = float(rx.get("optical_zone") or plan.get("optical_zone_mm") or 6.5)
                            depth = calculate_lenticule_thickness_smile(float(rx["sph"]), float(rx.get("cyl", 0)), oz)
                            len_th = str(int(depth + (float(min_th) if min_th not in (None, "") else 15.0)))
                    except: pass
                    c.drawString(300, y, f"Cap: {cap_th} / Min: {min_th} / Len: {len_th} µm")
                y -= 20
        
        c.setLineWidth(0.5)
        c.line(50, y+10, width-50, y+10)
        y -= 10

    c.save()
    return str(out_path)

def generate_patient_postop_pdf(row: Dict[str, Any], db: Dict[str, Any], tmp_dir: Path) -> Optional[str]:
    if not PDF_AVAILABLE: return None
    pid = str(row.get('id', 'unknown'))
    out_path = tmp_dir / f"patient_report_{pid}.pdf"
    c = canvas.Canvas(str(out_path), pagesize=A4)
    width, height = A4
    font_name = _register_cyrillic_font(tmp_dir)
    
    p = _get_p(db, pid)
    vid, m = _get_m(db, pid)
    f_db = db.get("forms", {}).get(pid, {})
    prim = f_db.get("primary", {})
    
    y = height - 50
    c.setFont(font_name, 18)
    c.drawString(50, y, f"Медицинская карта: {row.get('n', '-')}")
    y -= 25
    
    c.setFont(font_name, 11)
    c.drawString(50, y, f"ID: {pid} | Возраст: {row.get('a', '-')} | Пол: {str(prim.get('sex', '-')).upper()} | Телефон: {p.get('phone', '-')}")
    y -= 15
    c.drawString(50, y, f"Дата операции: {row.get('d', '-')} | Метод: {row.get('m', '-')}")
    y -= 25
    
    def safe_str(val): return str(val) if val not in (None, "") else "-"
    def fmt_rx(obj):
        if not isinstance(obj, dict): return "-"
        rx = obj.get('ref') or obj.get('bcva_rx') or obj.get('glasses_rx') or obj.get('glasses') or obj
        sph = rx.get('sph'); cyl = rx.get('cyl'); ax = rx.get('axis')
        va = obj.get('va') or obj.get('bcva') or obj.get('uva') or obj.get('ucva')
        if not va and isinstance(obj.get('glasses'), dict): va = obj['glasses'].get('va')
        if sph is None and cyl is None: return "-"
        try: s = f"{float(sph):+.2f}"
        except: s = str(sph)
        try: c = f"{float(cyl):+.2f}" if float(cyl) != 0 else "0.00"
        except: c = str(cyl)
        try: a = str(int(float(ax)))
        except: a = str(ax)
        res = f"{s} {c} x {a}"
        if va not in (None, ""): res += f" (VA: {va})"
        return res

    def draw_section(title, data_dict):
        nonlocal y
        if not any(str(od_v).replace(" (VA: -)", "") != "-" or str(os_v).replace(" (VA: -)", "") != "-" for _, od_v, os_v in data_dict): return
        if y < 100: c.showPage(); y = height - 50
        c.setFont(font_name, 12); c.setFillColorRGB(0.14, 0.5, 0.8); c.drawString(50, y, title); c.setFillColorRGB(0, 0, 0); y -= 15
        c.setFont(font_name, 10); c.setFillColorRGB(0.5, 0.5, 0.5); c.drawString(150, y, "OD (Правый)"); c.drawString(350, y, "OS (Левый)"); c.setFillColorRGB(0, 0, 0); y -= 15
        for label, val_od, val_os in data_dict:
            if str(val_od).replace(" (VA: -)", "") == "-" and str(val_os).replace(" (VA: -)", "") == "-": continue
            c.drawString(50, y, label); c.drawString(150, y, str(val_od).replace(" (VA: -)", "")); c.drawString(350, y, str(val_os).replace(" (VA: -)", "")); y -= 15
        y -= 5

    # ... Здесь мы рисуем блоки (Манифест, План и т.д.), как было в основном файле ...
    def fmt_iol(side):
        if not m.get("iol_calc"): return "-"
        iol = m["iol_calc"]
        sel = iol.get("selected_iol", {}).get(side, {})
        if sel and sel.get("power"):
            return f"{sel['power']} D (Tgt: {sel.get('target','-')})"
        if iol.get("barrett", {}).get(side):
            return f"Эмм: {iol['barrett'][side].get('p_emmetropia', 0):.2f} D"
        elif iol.get("kane", {}).get(side):
            return f"Эмм: {iol['kane'][side].get('p_emmetropia', 0):.2f} D (Kane)"
        return "-"
        
    if m.get("iol_calc"):
        draw_section("Сводка Катаракта", [("Ожид. Эмметропия", fmt_iol("od"), fmt_iol("os")), ("Линза", m["iol_calc"].get("lens_name", "-"), m["iol_calc"].get("lens_name", "-"))])
    else:
        draw_section("Сводка", [("План Rx", fmt_rx(m.get('surgery_plan',{}).get('od')), fmt_rx(m.get('surgery_plan',{}).get('os')))])

    if y < 150: c.showPage(); y = height - 50
    c.setFont(font_name, 12); c.setFillColorRGB(0.14, 0.5, 0.8); c.drawString(50, y, "Динамика (Post-Op)"); c.setFillColorRGB(0, 0, 0); y -= 20
    c.setFont(font_name, 10)
    history = sorted(row.get("h", []), key=lambda x: x.get("d", ""), reverse=True)
    if not history: c.drawString(50, y, "Нет данных.")
    else:
        def _fmt_po(r):
            if not r: return "-"
            s = f"{float(r.get('s',0)):+.2f} {float(r.get('c',0)):+.2f} x {int(r.get('a',0))}" if r.get('s') is not None else ""
            if r.get('v'): s += f" (VA {r.get('v')})"
            return s if s else "-"
        for h in history:
            if y < 100: c.showPage(); y = height - 50
            c.setFont(font_name, 11); c.setFillColorRGB(0.1, 0.6, 0.3); c.drawString(50, y, f"📅 {h.get('p', '-')} ({h.get('d', '-')})"); c.setFillColorRGB(0, 0, 0); y -= 15
            c.setFont(font_name, 10)
            if h.get("nt"): c.drawString(60, y, f"📝 Заметки: {h.get('nt')}"); y -= 15
            c.drawString(60, y, "OD:"); c.drawString(85, y, _fmt_po(h.get('od')))
            c.drawString(250, y, "OS:"); c.drawString(275, y, _fmt_po(h.get('os')))
            if h.get("ou") and h.get("ou").get("v"): c.drawString(420, y, f"OU: VA {h.get('ou').get('v')}")
            y -= 20
    c.save()
    return str(out_path)

def generate_outcomes_pdf(data: List[Dict[str, Any]], tmp_dir: Path) -> Optional[str]:
    if not PDF_AVAILABLE: return None
    out_path = tmp_dir / f"outcomes_{int(time.time())}.pdf"
    c = canvas.Canvas(str(out_path), pagesize=A4)
    width, height = A4
    font_name = _register_cyrillic_font(tmp_dir)
    
    y = height - 50
    c.setFont(font_name, 16)
    c.drawString(50, y, f"Сводный отчет: Результаты операций ({datetime.date.today()})")
    y -= 30
    
    for row in data:
        if y < 100:
            c.showPage()
            y = height - 50
        c.setFont(font_name, 12)
        c.drawString(50, y, f"{row.get('n')} ({row.get('d')}) - {row.get('m')}")
        y -= 15
        c.setFont(font_name, 10)
        def _fmt(r): return f"{r.get('s',0):+.2f} {r.get('c',0):+.2f} x {r.get('a',0)}" + (f" (VA {r.get('v')})" if r.get('v') else "") if r else "-"
        c.drawString(50, y, f"OD  Pre: {_fmt(row.get('od_pre'))}"); c.drawString(280, y, f"Plan: {_fmt(row.get('od_pln'))}"); c.drawString(450, y, f"Post: {_fmt(row.get('od_pst'))}")
        y -= 12
        c.drawString(50, y, f"OS  Pre: {_fmt(row.get('os_pre'))}"); c.drawString(280, y, f"Plan: {_fmt(row.get('os_pln'))}"); c.drawString(450, y, f"Post: {_fmt(row.get('os_pst'))}")
        y -= 25
        
    c.save()
    return str(out_path)

def generate_postop_excel_bytes(post_op_db: Dict[str, Any], db: Dict[str, Any]) -> Optional[io.BytesIO]:
    if not OPENPYXL_AVAILABLE: return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    ws.append(["ID", "ФИО", "Дата операции", "Дата осмотра", "Срок", "Результат (Summary)", "VA OD", "VA OS"])
    for cell in ws[1]: cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center")
        
    all_rows = []
    for pid, entries in post_op_db.items():
        p = _get_p(db, pid)
        if not p: continue
        op_date = db.get("forms", {}).get(pid, {}).get("op_date", "-")
        
        for r in entries:
            summ = r.get("summary", "-")
            data_struct = r.get("data")
            va_od = va_os = None
            if isinstance(data_struct, dict):
                try: va_od = float(str(data_struct.get("od", {}).get("va", "")).replace(',', '.'))
                except: pass
                try: va_os = float(str(data_struct.get("os", {}).get("va", "")).replace(',', '.'))
                except: pass
            if va_od is None:
                m_od = re.search(r'(?:OD|O\.D\.|Прав\.?)\s*[:=]?\s*([0-1](?:[\.,]\d+)?)', summ, re.I)
                if m_od:
                    try: va_od = float(m_od.group(1).replace(',', '.'))
                    except: pass
            if va_os is None:
                m_os = re.search(r'(?:OS|O\.S\.|Лев\.?)\s*[:=]?\s*([0-1](?:[\.,]\d+)?)', summ, re.I)
                if m_os:
                    try: va_os = float(m_os.group(1).replace(',', '.'))
                    except: pass
            all_rows.append([pid, p.get("name", "-"), op_date, r.get("date", "-"), r.get("period", "-"), summ, va_od, va_os])
    
    all_rows.sort(key=lambda x: x[3], reverse=True)
    for row in all_rows: ws.append(row)
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 20

    if len(all_rows) > 0:
        chart = BarChart()
        chart.title = "Острота зрения (последние осмотры)"
        limit = min(len(all_rows), 15)
        data = Reference(ws, min_col=7, min_row=1, max_row=limit+1, max_col=8)
        cats = Reference(ws, min_col=2, min_row=2, max_row=limit+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "J2")

    # Векторный анализ
    ws_vec = wb.create_sheet("Векторный анализ")
    legend_font = Font(bold=True, color="4F81BD")
    for row in [
        ["Показатель", "Описание и Идеальное значение"],
        ["TIA", "Цель (Target). Вектор астигматизма, который планировалось исправить."],
        ["SIA", "Факт (Surgically Induced). Вектор, который фактически убрал лазер."],
        ["DV", "Остаток (Difference Vector). ИДЕАЛ = 0.00 D. Разница между фактом и целью."],
        ["ME", "Ошибка величины (Magnitude of Error). ИДЕАЛ = 0.00 D. Отрицательное = недокоррекция."],
        ["AE", "Ошибка угла (Angle of Error). ИДЕАЛ = 0°. Угол промаха оси (показывает циклоторсию)."],
        ["CI", "Индекс коррекции (Correction Index). ИДЕАЛ = 1.0. CI < 1.0 = недокоррекция (надо добавить %)."]
    ]:
        ws_vec.append(row)
        ws_vec.cell(row=ws_vec.max_row, column=1).font = legend_font
        ws_vec.cell(row=ws_vec.max_row, column=2).font = Font(italic=True)
        
    ws_vec.append([])
    ws_vec.append(["ID", "ФИО", "Eye", "Laser", "Period", "Pre_Cyl", "Pre_Ax", "Post_Cyl", "Post_Ax", "TIA", "TIA_Ax", "SIA", "SIA_Ax", "DV", "DV_Ax", "ME", "AE", "CI"])
    
    header_fill_vec = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    for cell in ws_vec[ws_vec.max_row]:
        cell.font = header_font; cell.fill = header_fill_vec; cell.alignment = Alignment(horizontal="center")
        
    for pid, entries in post_op_db.items():
        p = _get_p(db, pid)
        if not p: continue
        vid, m = _get_m(db, pid)
        plan = m.get("surgery_plan", {})
        man = m.get("manifest", {})
        for r in entries:
            post_data = r.get("data", {})
            if not isinstance(post_data, dict): continue
            for side in ["od", "os"]:
                pst = post_data.get(side)
                if not isinstance(pst, dict): continue
                eye_man = man.get(side, {})
                rx = eye_man.get("bcva_rx") if isinstance(eye_man.get("bcva_rx"), dict) else eye_man
                try:
                    pre_c = float(str(rx.get("cyl")).replace(',','.')); pre_a = float(str(rx.get("axis")).replace(',','.'))
                    post_c = float(str(pst.get("cyl")).replace(',','.')); post_a = float(str(pst.get("ax")).replace(',','.'))
                    vec = alpins_vector_analysis(pre_c, pre_a, post_c, post_a)
                    ws_vec.append([pid, p.get("name", "-"), side.upper(), plan.get("laser_type", ""), r.get("period", "-"), pre_c, int(pre_a), post_c, int(post_a), vec["TIA"], vec["TIA_ax"], vec["SIA"], vec["SIA_ax"], vec["DV"], vec["DV_ax"], vec["ME"], vec["AE"], vec["CI"]])
                except Exception: pass
    for col in ws_vec.columns: ws_vec.column_dimensions[col[0].column_letter].width = 10
    ws_vec.column_dimensions['B'].width = 85

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out